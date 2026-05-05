from io import BytesIO

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app import crud, schemas
from app.auth import obter_usuario_atual
from app.database import get_db
from app.services.importacao_excel import importar_headcount_excel

router = APIRouter(dependencies=[Depends(obter_usuario_atual)])


@router.get("/headcount/modelo-excel")
def baixar_modelo_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "headcount"
    colunas = [
        "tipo",
        "matricula",
        "nome",
        "cargo",
        "empresa",
        "departamento",
        "centro_custo",
        "grupo",
        "salario",
        "status",
        "data_entrada",
        "data_saida_prevista",
        "dependente_1_nome",
        "dependente_1_idade",
        "dependente_2_nome",
        "dependente_2_idade",
        "dependente_3_nome",
        "dependente_3_idade",
        "justificativa",
    ]
    ws.append(colunas)
    ws.append([
        "colaborador",
        "001",
        "Ana Souza",
        "Analista",
        "CLI",
        "RH",
        "CC-RH",
        "Geral",
        4200,
        "ativo",
        "01/01/2026",
        "",
        "Lucas Souza",
        8,
        "Marina Souza",
        12,
        "",
        "",
        "Carga inicial do budget",
    ])
    ws.append([
        "vaga",
        "V001",
        "Vaga Analista Operacoes",
        "Analista",
        "CLI",
        "Operacoes",
        "CC-OPS",
        "Operacional",
        3800,
        "planejado",
        "01/03/2026",
        "",
        "Dependente previsto 1",
        "",
        "Dependente previsto 2",
        "",
        "Dependente previsto 3",
        "",
        "Nova vaga aprovada",
    ])

    orientacao = wb.create_sheet("arvore_importacao")
    orientacao.append(["Nivel", "Campo", "Obrigatorio", "Uso no sistema"])
    orientacoes = [
        ["1 - Versao", "versao_id", "Sim", "Escolhida na tela antes da importacao; separa um cenario do outro."],
        ["2 - Empresa", "empresa", "Recomendado", "Base para filtros, indicadores e consolidacao por empresa."],
        ["3 - Departamento", "departamento", "Sim", "Base para agrupamentos, relatorios e indicadores por area."],
        ["4 - Centro de custo", "centro_custo", "Sim", "Base para rateio e visao financeira."],
        ["5 - Cargo", "cargo", "Sim", "Base para regras de verba, PLR, adicionais e indicadores."],
        ["6 - Grupo", "grupo", "Recomendado", "Liga a pessoa ou vaga a reajustes e premissas especificas."],
        ["7 - Titular/Vaga", "tipo, matricula, nome, status", "Sim", "Define se conta como colaborador existente ou vaga planejada."],
        ["8 - Periodo", "data_entrada, data_saida_prevista", "Recomendado", "Define em quais meses o headcount entra no calculo."],
        ["9 - Remuneracao", "salario", "Sim", "Base para salario, provisoes, encargos e beneficios dependentes de salario."],
        ["10 - Dependentes", "dependente_1_nome...dependente_3_idade", "Opcional", "Base para plano de saude, odontologico e beneficios por vida/faixa etaria."],
        ["11 - Auditoria", "justificativa", "Sim", "Explica por que o registro entrou ou foi alterado no budget."],
    ]
    for linha in orientacoes:
        orientacao.append(linha)

    header_fill = PatternFill("solid", fgColor="00613A")
    for sheet in [ws, orientacao]:
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for column in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max_len + 3, 52)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": "attachment; filename=modelo_headcount.xlsx"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/versoes/{versao_id}/headcount", response_model=list[schemas.HeadcountOut])
def listar(versao_id: int, db: Session = Depends(get_db)):
    return crud.listar_headcount(db, versao_id)


@router.post("/versoes/{versao_id}/headcount", response_model=schemas.HeadcountOut)
def criar(versao_id: int, dados: schemas.HeadcountCreate, db: Session = Depends(get_db)):
    return crud.criar_headcount(db, versao_id, dados)


@router.put("/headcount/{item_id}", response_model=schemas.HeadcountOut)
def atualizar(item_id: int, dados: schemas.HeadcountUpdate, db: Session = Depends(get_db)):
    return crud.atualizar_headcount(db, item_id, dados)


@router.delete("/headcount/{item_id}")
def excluir(item_id: int, db: Session = Depends(get_db)):
    crud.excluir_headcount(db, item_id)
    return {"ok": True}


@router.post("/versoes/{versao_id}/headcount/importar")
def importar(versao_id: int, arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    return importar_headcount_excel(db, versao_id, arquivo)

